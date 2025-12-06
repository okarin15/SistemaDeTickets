from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.views.decorators.http import require_POST
from django.contrib import messages
from .models import Ticket, Categoria, UserProfile, Comentario, HistorialTicket, FAQ, Area
import json
from django.urls import reverse
from django.core.serializers.json import DjangoJSONEncoder
from django.utils import timezone
from django.db.models import Count, Q

# --- IMPORTS PARA EXPORTACIÓN ---
from django.http import HttpResponse
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# Pega esta función auxiliar
def registrar_historial(ticket, usuario, accion):
    HistorialTicket.objects.create(
        ticket=ticket,
        usuario=usuario,
        accion=accion,
        fecha=timezone.now()
    )


def login_view(request):
    # Si ya está logueado, redirigir
    if request.user.is_authenticated:
        if request.user.is_staff:
            return redirect('dashboard_admin')
        # Verificar si ya tiene rol asignado
        try:
            if hasattr(request.user, 'userprofile'):
                if request.user.userprofile.role == 'tech':
                    return redirect('dashboard_tecnico')
        except:
            pass
        return redirect('dashboard_user')
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)
            
            # --- ZONA DE DIAGNÓSTICO (MIRA TU TERMINAL) ---
            print(f"--> USUARIO LOGUEADO: {user.username}")
            
            if user.is_staff:
                print("--> ES STAFF: Redirigiendo a Admin")
                return redirect('dashboard_admin')
            
            try:
                if hasattr(user, 'userprofile'):
                    rol_actual = user.userprofile.role
                    print(f"--> ROL ENCONTRADO: {rol_actual}")
                    
                    if rol_actual == 'tech':
                        print("--> ROL ES TECH: Redirigiendo a Dashboard Técnico")
                        return redirect('dashboard_tecnico')
                    else:
                        print(f"--> ROL NO ES TECH (es {rol_actual}): Redirigiendo a Usuario")
                else:
                    print("--> ERROR: El usuario NO tiene perfil creado")
            except Exception as e:
                print(f"--> EXCEPCIÓN: {e}")

            return redirect('dashboard_user')
        else:
            messages.error(request, 'Usuario o contraseña incorrectos')
    
    return render(request, 'login.html')

@login_required
def logout_view(request):
    logout(request)
    messages.success(request, 'Sesión cerrada correctamente')
    return redirect('login')

@login_required
def dashboard_admin(request):
    # Validar que sea admin (Staff)
    if not request.user.is_staff:
        messages.error(request, 'Acceso denegado. No eres administrador.')
        return redirect('dashboard_user')

    # 1. OBTENER USUARIOS DE LA BD REAL
    # 1. USUARIOS
    users_qs = User.objects.select_related('userprofile').all()
    users_data = []
    
    for u in users_qs:
        rol = 'user'
        try:
            if hasattr(u, 'userprofile'):
                rol = u.userprofile.role
        except:
            pass

        users_data.append({
            'id': u.id,
            'username': u.username,
            'nombre': f"{u.first_name} {u.last_name}" if u.first_name else u.username,
            'email': u.email,
            'rol': rol,
            'estado': 'active' if u.is_active else 'inactive',
            'fecha_registro': u.date_joined.strftime('%Y-%m-%d'),
            'ultimo_acceso': u.last_login.strftime('%Y-%m-%d') if u.last_login else '-'
        })

    # 2. OBTENER TICKETS DE LA BD REAL
    # 2. TICKETS
    tickets_qs = Ticket.objects.all().order_by('-fecha_creacion')
    tickets_data = []

    for t in tickets_qs:
        # --- CÁLCULO DE TIEMPO REAL ---
        horas_resolucion = 0
        if t.fecha_cierre and t.fecha_creacion:
            diferencia = t.fecha_cierre - t.fecha_creacion
            # Convertimos a horas (total_seconds / 3600)
            horas_resolucion = round(diferencia.total_seconds() / 3600, 1)
        
        # --- CÁLCULO DE SLA ---
        import datetime
        horas_limite = {'critical': 4, 'high': 24, 'medium': 48, 'low': 72}.get(t.prioridad, 48)
        fecha_fin = t.fecha_cierre if t.fecha_cierre else timezone.now()
        horas_transcurridas = (fecha_fin - t.fecha_creacion).total_seconds() / 3600
        
        if horas_transcurridas > horas_limite: sla_status = 'overdue'
        elif horas_transcurridas > (horas_limite * 0.8): sla_status = 'warning'
        else: sla_status = 'ok'
        
        # ------------------------------
        # 1. Recuperar el historial de ese ticket
        historial_list = []
        for h in t.historial.all().order_by('-fecha'):
            historial_list.append({
                'usuario': h.usuario.username,
                'accion': h.accion,
                'fecha': h.fecha.strftime('%d/%m/%Y %H:%M')
            })

        # 2. Agregarlo al diccionario del ticket
        tickets_data.append({
            'id': t.id,
            'titulo': t.titulo,
            'categoria': t.categoria.nombre if t.categoria else "Sin Categoría", # Asegúrate de que coincida con tu código anterior
            'area': t.area.nombre if t.area else "No especificada",
            'usuario': t.solicitante.username,
            'prioridad': t.prioridad, 
            'estado': t.estado,       
            'fecha': t.fecha_creacion.strftime('%d/%m/%Y'),
            
            # --- NUEVOS DATOS PARA EL REPORTE ---
            'horas': horas_resolucion, 
            'cerrado': True if t.fecha_cierre else False,
            'calificacion': t.calificacion,
            'sla': sla_status,
            'horas_limite': horas_limite,
            # ------------------------------------
            
            'agente': t.asignado_a.username if t.asignado_a else "",
            'descripcion': t.descripcion,
        })

    # 3. CATEGORÍAS (NUEVO)
    # Contamos cuántos tickets tiene cada categoría para mostrarlo en la tabla
    categorias_qs = Categoria.objects.all()
    categorias_data = []
    for c in categorias_qs:
        cantidad = Ticket.objects.filter(categoria=c).count()
        categorias_data.append({
            'id': c.id,
            'nombre': c.nombre,
            'descripcion': c.descripcion,
            'activo': c.activo,
            'cantidad_tickets': cantidad
        })

    # --- OBTENER FAQs (Poner esto en dashboard_user, dashboard_admin y dashboard_tecnico) ---
    faqs_qs = FAQ.objects.filter(activo=True).order_by('-fecha_creacion')
    faqs_data = []
    for f in faqs_qs:
        faqs_data.append({
            'id': f.id,
            'pregunta': f.pregunta,
            'respuesta': f.respuesta
        })
    # -------------------------------------------------------------------------------------

    # 4. ÁREAS
    areas_qs = Area.objects.filter(activo=True)
    areas_data = [{'id': a.id, 'nombre': a.nombre, 'descripcion': a.descripcion} for a in areas_qs]


    context = {
        'users_json': json.dumps(users_data, cls=DjangoJSONEncoder),
        'tickets_json': json.dumps(tickets_data, cls=DjangoJSONEncoder),
        'categorias_json': json.dumps(categorias_data, cls=DjangoJSONEncoder), 
        'faqs_json': json.dumps(faqs_data, cls=DjangoJSONEncoder), # <--- NUEVA LÍNEA
        'areas_json': json.dumps(areas_data, cls=DjangoJSONEncoder),
    }

    return render(request, 'dashboard_admin.html', context)

@login_required
def dashboard_user(request):
    # --- BLOQUE DE SEGURIDAD (NUEVO) ---
    # 1. Si es Admin (Staff), lo mandamos a su panel
    if request.user.is_staff:
        return redirect('dashboard_admin')

    # 2. Si es Técnico, lo mandamos a su panel
    try:
        if hasattr(request.user, 'userprofile') and request.user.userprofile.role == 'tech':
            return redirect('dashboard_tecnico')
    except:
        pass
    # -----------------------------------

    # 1. Obtener tickets reales del usuario
    user_tickets = Ticket.objects.filter(solicitante=request.user).order_by('-fecha_creacion')
    
    # 2. Serializar para JavaScript
    tickets_data = []
    for t in user_tickets:
        comentarios_list = []
        for c in t.comentarios.all():
            comentarios_list.append({
                'autor': c.autor.username, # O c.autor.first_name
                'fecha': c.fecha.strftime('%d/%m/%Y %H:%M'),
                'contenido': c.contenido
            })

        tickets_data.append({
            'id': t.id,
            'titulo': t.titulo,
            'categoria': t.categoria.nombre if t.categoria else "Sin Categoría",
            'area': t.area.nombre if t.area else "No especificada",
            'prioridad': t.prioridad, 
            'estado': t.estado,       
            'fecha_creacion': t.fecha_creacion.strftime('%Y-%m-%d %H:%M'),
            'ultima_actualizacion': t.fecha_actualizacion.strftime('%Y-%m-%d %H:%M'),
            'descripcion': t.descripcion,
            'archivo_url': t.archivo.url if t.archivo else None, 
            'comentarios': comentarios_list,
            'calificacion': t.calificacion, # <--- AGREGAR ESTA LÍNEA
        })

    # 3. Categorías
    categorias = Categoria.objects.filter(activo=True)
    areas = Area.objects.filter(activo=True)
    
    # --- OBTENER FAQs (Poner esto en dashboard_user, dashboard_admin y dashboard_tecnico) ---
    faqs_qs = FAQ.objects.filter(activo=True).order_by('-fecha_creacion')
    faqs_data = []
    for f in faqs_qs:
        faqs_data.append({
            'id': f.id,
            'pregunta': f.pregunta,
            'respuesta': f.respuesta
        })
    # -------------------------------------------------------------------------------------
    
    context = {
        'tickets_json': json.dumps(tickets_data, cls=DjangoJSONEncoder),
        'categorias': categorias,
        'areas': areas,
        'faqs_json': json.dumps(faqs_data, cls=DjangoJSONEncoder), # <--- NUEVA LÍNEA
    }
    return render(request, 'dashboard_user.html', context)

@login_required
def crear_ticket(request):
    # Obtener categorías
    categorias = Categoria.objects.filter(activo=True)
    # --- NUEVO: OBTENER ÁREAS ---
    areas = Area.objects.filter(activo=True)

    if request.method == 'POST':
        titulo = request.POST.get('titulo')
        categoria_id = request.POST.get('categoria')
        descripcion = request.POST.get('descripcion')
        # --- NUEVO: CAPTURAR EL ÁREA ---
        area = request.POST.get('area')
        area_obj = get_object_or_404(Area, id=area) if area else None
        
        # Capturar el archivo
        archivo = request.FILES.get('archivo')

        prioridad_defecto = 'medium' 

        # Validar categoría
        categoria = get_object_or_404(Categoria, id=categoria_id)

        # Crear ticket en BD
        nuevo_ticket = Ticket.objects.create(
            titulo=titulo,
            categoria=categoria,
            area=area_obj,  # <--- GUARDARLO AQUÍ
            prioridad=prioridad_defecto,
            descripcion=descripcion,
            solicitante=request.user,
            estado='new',
            archivo=archivo
        )
        
        # Guardar historial
        registrar_historial(nuevo_ticket, request.user, "Creación del Ticket")

        # --- CÓDIGO NUEVO: ENVIAR CORREO REAL ---
        # Importamos aquí mismo para asegurar que no falte arriba
        from django.core.mail import send_mail
        from django.conf import settings
        
        try:
            asunto = f'Ticket Creado: T-{nuevo_ticket.id:04d} - {nuevo_ticket.titulo}'
            
            # --- MENSAJE CORREGIDO (SIN PRIORIDAD) ---
            mensaje = f"""
            Hola {request.user.username},
            
            Tu solicitud ha sido recibida correctamente.
            
            ID: T-{nuevo_ticket.id:04d}
            Título: {nuevo_ticket.titulo}
            
            Un técnico revisará tu caso pronto.
            
            Atte,
            Equipo de Soporte Coyahue
            """
            
            email_destino = request.user.email
            
            if email_destino:
                send_mail(
                    asunto,
                    mensaje,
                    settings.EMAIL_HOST_USER, # Remitente
                    [email_destino],          # Destinatario
                    fail_silently=False,
                )
                print(f"--> Correo enviado a {email_destino}")
            else:
                print("--> El usuario no tiene email configurado.")
                
        except Exception as e:
            print(f"--> Error enviando correo: {e}") 
        # ----------------------------------------
        
        messages.success(request, 'Ticket creado exitosamente')
        return redirect('dashboard_user') 
    
    return render(request, 'crear_ticket.html', {'categorias': categorias, 'areas': areas})

@login_required
def mis_tickets(request):
    user_tickets = Ticket.objects.filter(solicitante=request.user).order_by('-fecha_creacion')
    return render(request, 'mis_tickets.html', {'tickets': user_tickets})

@login_required
def actualizar_ticket(request, ticket_id):
    # Solo staff puede actualizar estado por esta vía rápida
    if not request.user.is_staff:
        messages.error(request, 'No tienes permisos para actualizar tickets')
        return redirect('dashboard_user')
    
    ticket = get_object_or_404(Ticket, id=ticket_id)
    
    if request.method == 'POST':
        nuevo_estado = request.POST.get('estado')
        if nuevo_estado:
            ticket.estado = nuevo_estado
            ticket.save()
            messages.success(request, f'Ticket T-{ticket.id:04d} actualizado')
    
    return redirect('dashboard_admin')

@login_required
@require_POST
def guardar_categoria(request):
    # Solo admin puede hacer esto
    if not request.user.is_staff:
        messages.error(request, 'No tienes permisos.')
        return redirect('dashboard_user')

    try:
        # Obtener datos del formulario
        cat_id = request.POST.get('cat_id')
        nombre = request.POST.get('nombre')
        descripcion = request.POST.get('descripcion')
        activo_str = request.POST.get('activo')
        
        is_active = True if activo_str == 'true' else False

        if cat_id:
            # --- EDITAR EXISTING ---
            categoria = get_object_or_404(Categoria, id=cat_id)
            categoria.nombre = nombre
            categoria.descripcion = descripcion
            categoria.activo = is_active
            categoria.save()
            messages.success(request, f'Categoría "{nombre}" actualizada correctamente.')
        else:
            # --- CREAR NUEVA ---
            Categoria.objects.create(
                nombre=nombre,
                descripcion=descripcion,
                activo=is_active
            )
            messages.success(request, f'Categoría "{nombre}" creada correctamente.')

    except Exception as e:
        messages.error(request, f'Error al guardar: {str(e)}')

    # Volver al dashboard
    # CAMBIA EL FINAL POR ESTO:
    response = redirect('dashboard_admin')
    response['Location'] += '?tab=categorias' # <--- Truco para volver a la pestaña
    return response

@login_required
@require_POST
def guardar_usuario(request):
    # Seguridad: Solo admin puede entrar aquí
    if not request.user.is_staff:
        messages.error(request, 'No tienes permisos para gestionar usuarios.')
        return redirect('dashboard_user')

    try:
        # Recibir datos del formulario
        user_id = request.POST.get('user_id')
        username = request.POST.get('username')
        email = request.POST.get('email')
        rol = request.POST.get('rol')
        password = request.POST.get('password')
        nombre_completo = request.POST.get('nombre')

        if user_id:
            # === EDITAR USUARIO EXISTENTE ===
            usuario = get_object_or_404(User, id=user_id)
            usuario.username = username
            usuario.email = email
            
            # Separar nombre y apellido
            if nombre_completo:
                partes = nombre_completo.split(' ', 1)
                usuario.first_name = partes[0]
                usuario.last_name = partes[1] if len(partes) > 1 else ''
            
            # Solo cambiamos la contraseña si escribieron algo
            if password:
                usuario.set_password(password)
            
            usuario.save()

            # Actualizar el Rol (Perfil)
            if hasattr(usuario, 'userprofile'):
                usuario.userprofile.role = rol
                usuario.userprofile.save()
            
            messages.success(request, f'Usuario "{username}" actualizado correctamente.')

        else:
            # === CREAR NUEVO USUARIO ===
            if User.objects.filter(username=username).exists():
                messages.error(request, 'El nombre de usuario ya existe.')
                return redirect('dashboard_admin')

            # Creamos el usuario (la contraseña se encripta aquí)
            nuevo_usuario = User.objects.create_user(username=username, email=email, password=password)
            
            if nombre_completo:
                partes = nombre_completo.split(' ', 1)
                nuevo_usuario.first_name = partes[0]
                nuevo_usuario.last_name = partes[1] if len(partes) > 1 else ''
            
            nuevo_usuario.save()

            # Actualizamos el rol (El perfil ya se creó automáticamente por los Signals)
            if hasattr(nuevo_usuario, 'userprofile'):
                nuevo_usuario.userprofile.role = rol
                nuevo_usuario.userprofile.save()

            messages.success(request, f'Usuario "{username}" creado correctamente.')

    except Exception as e:
        messages.error(request, f'Error al procesar: {str(e)}')

    # Volver a la pestaña de usuarios
    response = redirect('dashboard_admin')
    response['Location'] += '?tab=usuarios'
    return response

@login_required
@require_POST
def agregar_comentario(request, ticket_id):
    ticket = get_object_or_404(Ticket, id=ticket_id)
    contenido = request.POST.get('contenido')

    if contenido:
        Comentario.objects.create(
            ticket=ticket,
            autor=request.user,
            contenido=contenido
        )
        
        # --- NUEVO: NOTIFICACIÓN DE COMENTARIO ---
        try:
            from django.core.mail import send_mail
            from django.conf import settings
            
            destinatario_email = None
            nombre_destinatario = ""
            
            # LÓGICA INTELIGENTE:
            # Si quien comenta NO es el dueño del ticket (ej: es el técnico), avisamos al dueño
            if request.user != ticket.solicitante:
                destinatario_email = ticket.solicitante.email
                nombre_destinatario = ticket.solicitante.username
                tipo_aviso = "El técnico ha respondido a tu ticket"
                
            # Si quien comenta ES el dueño, avisamos al técnico asignado (si existe)
            elif ticket.asignado_a and ticket.asignado_a.email:
                destinatario_email = ticket.asignado_a.email
                nombre_destinatario = ticket.asignado_a.username
                tipo_aviso = "El usuario respondió al ticket"

            if destinatario_email:
                asunto = f'Nuevo Mensaje en Ticket T-{ticket.id:04d}'
                mensaje = f"""
                Hola {nombre_destinatario},
                
                {tipo_aviso} "{ticket.titulo}".
                
                Comentario:
                "{contenido}"
                
                Ingresa al portal para responder.
                """
                send_mail(asunto, mensaje, settings.EMAIL_HOST_USER, [destinatario_email], fail_silently=True)
                
        except Exception as e:
            print(f"Error enviando correo comentario: {e}")
        # -----------------------------------------

        messages.success(request, 'Comentario agregado correctamente')
    
    # Redirección inteligente
    es_tecnico = request.user.is_staff
    try:
        if hasattr(request.user, 'userprofile') and request.user.userprofile.role == 'tech':
            es_tecnico = True
    except:
        pass

    if es_tecnico:
        return redirect('dashboard_tecnico')
    else:
        return redirect('dashboard_user')

@login_required
def dashboard_tecnico(request):
    # Validar permiso (staff o rol tecnico)
    # Por ahora dejamos pasar a staff o si tiene perfil tecnico
    es_tecnico = request.user.is_staff or (hasattr(request.user, 'userprofile') and request.user.userprofile.role == 'tech')
    
    if not es_tecnico:
        messages.error(request, 'No tienes permisos de técnico.')
        return redirect('dashboard_user')
    
    # --- CAMBIO AQUÍ: ELIMINAMOS EL FILTRO DE PRIVACIDAD ---
    # Antes: .filter(Q(asignado_a__isnull=True) | Q(asignado_a=request.user))
    # Ahora: Traemos TODOS los tickets para que puedan cubrirse entre compañeros
    all_tickets = Ticket.objects.select_related('solicitante', 'asignado_a', 'categoria').all().order_by('-fecha_creacion')

    # 2. Serializar datos
    tickets_data = []
    for t in all_tickets:
        # Obtener comentarios
        comentarios_list = []
        for c in t.comentarios.all():
            comentarios_list.append({
                'autor': c.autor.username,
                'fecha': c.fecha.strftime('%d/%m/%Y %H:%M'),
                'contenido': c.contenido
            })

        # --- CÁLCULO DE SLA ---
        import datetime
        horas_limite = {'critical': 4, 'high': 24, 'medium': 48, 'low': 72}.get(t.prioridad, 48)
        fecha_fin = t.fecha_cierre if t.fecha_cierre else timezone.now()
        horas_transcurridas = (fecha_fin - t.fecha_creacion).total_seconds() / 3600
        
        if horas_transcurridas > horas_limite: sla_status = 'overdue'
        elif horas_transcurridas > (horas_limite * 0.8): sla_status = 'warning'
        else: sla_status = 'ok'


        tickets_data.append({
            'id': t.id,
            'titulo': t.titulo,
            'usuario': t.solicitante.username,
            'area': t.area.nombre if t.area else "No especificada",
            'categoria': t.categoria.nombre if t.categoria else "General",
            'prioridad': t.prioridad, 
            'estado': t.estado,       
            'fecha': t.fecha_creacion.strftime('%d/%m/%Y'),
            'agente': t.asignado_a.username if t.asignado_a else "", # Importante para saber si está asignado
            'descripcion': t.descripcion,
            'comentarios': comentarios_list,
            'archivo_url': t.archivo.url if t.archivo else None, # --- LÍNEA AGREGADA/MODIFICADA ---
            'sla': sla_status,
            'horas_limite': horas_limite,
        })

    # --- OBTENER FAQs (Poner esto en dashboard_user, dashboard_admin y dashboard_tecnico) ---
    faqs_qs = FAQ.objects.filter(activo=True).order_by('-fecha_creacion')
    faqs_data = []
    for f in faqs_qs:
        faqs_data.append({
            'id': f.id,
            'pregunta': f.pregunta,
            'respuesta': f.respuesta
        })
    # -------------------------------------------------------------------------------------

    context = {
        'tickets_json': json.dumps(tickets_data, cls=DjangoJSONEncoder),
        'user': request.user, # Para saber quién es el técnico actual en JS
        'faqs_json': json.dumps(faqs_data, cls=DjangoJSONEncoder), # <--- NUEVA LÍNEA
    }
    return render(request, 'dashboard_tecnico.html', context)

# --- NUEVA FUNCIÓN PARA "TOMAR" TICKET ---
@login_required
def tomar_ticket(request, ticket_id):
    ticket = get_object_or_404(Ticket, id=ticket_id)
    
    # Guardamos quién lo tenía antes para el historial
    antiguo_tecnico = ticket.asignado_a.username if ticket.asignado_a else "Nadie"
    
    # Asignar al usuario actual
    ticket.asignado_a = request.user
    ticket.estado = 'in-progress' # Lo ponemos en progreso automáticamente
    ticket.save()
    
    # Historial claro
    registrar_historial(ticket, request.user, f"Tomó el ticket (Reasignado de: {antiguo_tecnico})")
    
    messages.success(request, f'Ticket T-{ticket.id:04d} ahora está asignado a ti.')
    return redirect(reverse('dashboard_tecnico') + '?tab=mis-tickets')

@login_required
def cambiar_estado_ticket(request, ticket_id):
    ticket = get_object_or_404(Ticket, id=ticket_id)
    
    # Verificar permisos
    es_tecnico = False
    try:
        if hasattr(request.user, 'userprofile') and request.user.userprofile.role == 'tech':
            es_tecnico = True
    except:
        pass

    if not (request.user.is_staff or es_tecnico):
        messages.error(request, 'No tienes permiso para modificar este ticket.')
        return redirect('dashboard_tecnico')
    estado_anterior = ticket.get_estado_display()
    
    # --- LÓGICA DE ESTADOS ---
    if ticket.estado == 'new' or ticket.estado == 'pending':
        ticket.estado = 'in-progress'
    elif ticket.estado == 'in-progress':
        ticket.estado = 'resolved'
    elif ticket.estado == 'resolved' or ticket.estado == 'completed':
        ticket.estado = 'closed'
        ticket.fecha_cierre = timezone.now()
    elif ticket.estado == 'closed':
        ticket.estado = 'in-progress'
        ticket.fecha_cierre = None
    
    ticket.save()
    
    registrar_historial(ticket, request.user, f"Cambio de estado de '{estado_anterior}' a '{ticket.get_estado_display()}'")
    
    # --- NUEVO: NOTIFICACIÓN DE CIERRE/RESOLUCIÓN ---
    # Si el estado nuevo es Resuelto o Cerrado, avisamos al usuario
    if ticket.estado in ['resolved', 'completed', 'closed']:
        try:
            from django.core.mail import send_mail
            from django.conf import settings
            
            asunto = f'Actualización Ticket T-{ticket.id:04d}: {ticket.get_estado_display()}'
            mensaje = f"""
            Hola {ticket.solicitante.username},
            
            Tu ticket "{ticket.titulo}" ha cambiado de estado a: {ticket.get_estado_display().upper()}.
            
            Si tu problema ha sido solucionado, no necesitas hacer nada más.
            Si el problema persiste, por favor entra al sistema y agrega un comentario.
            
            Atte,
            Soporte Coyahue
            """
            
            if ticket.solicitante.email:
                send_mail(asunto, mensaje, settings.EMAIL_HOST_USER, [ticket.solicitante.email], fail_silently=True)
        except Exception:
            pass # No detenemos el flujo si falla el correo
    # -----------------------------------------------

    messages.success(request, f'Estado actualizado a: {ticket.get_estado_display()}')
    return redirect('dashboard_tecnico')

@login_required
def imprimir_ticket(request, ticket_id):
    ticket = get_object_or_404(Ticket, id=ticket_id)
    
    # --- VALIDACIÓN ESTRICTA ---
    # Permiso solo si:
    # 1. Es Admin (Staff)
    # 2. Es el Técnico ASIGNADO ESPECÍFICAMENTE a este ticket
    # 3. Es el Usuario dueño del ticket (Solicitante)
    
    tiene_permiso = (
        request.user.is_staff or 
        ticket.asignado_a == request.user or 
        ticket.solicitante == request.user
    )

    if not tiene_permiso:
        messages.error(request, 'No tienes permiso para imprimir este ticket.')
        return redirect('dashboard_user') # O dashboard_tecnico según corresponda

    return render(request, 'ticket_print.html', {'ticket': ticket})

@login_required
@require_POST
def calificar_ticket(request, ticket_id):
    ticket = get_object_or_404(Ticket, id=ticket_id)
    
    # Solo el dueño puede calificar
    if ticket.solicitante != request.user:
        messages.error(request, 'No puedes calificar este ticket.')
        return redirect('dashboard_user')

    calificacion = request.POST.get('calificacion')
    comentario = request.POST.get('comentario', '')

    if calificacion:
        ticket.calificacion = int(calificacion)
        ticket.comentario_calificacion = comentario
        
        # Opcional: Si califican, lo pasamos a Cerrado automáticamente si no lo estaba
        if ticket.estado == 'resolved':
            ticket.estado = 'closed'
            
        ticket.save()
        
        registrar_historial(ticket, request.user, f"Calificó el servicio con {calificacion} estrellas")
        messages.success(request, '¡Gracias por tu calificación!')

    return redirect('dashboard_user')

@login_required
@require_POST
def guardar_faq(request):
    # Validar permisos (Admin o Técnico)
    es_admin = request.user.is_staff
    es_tecnico = False
    try:
        if hasattr(request.user, 'userprofile') and request.user.userprofile.role == 'tech':
            es_tecnico = True
    except:
        pass

    if not (es_admin or es_tecnico):
        messages.error(request, 'No tienes permiso para gestionar FAQs.')
        return redirect('dashboard_user')

    faq_id = request.POST.get('faq_id')
    pregunta = request.POST.get('pregunta')
    respuesta = request.POST.get('respuesta')

    if faq_id:
        # Editar
        faq = get_object_or_404(FAQ, id=faq_id)
        faq.pregunta = pregunta
        faq.respuesta = respuesta
        faq.save()
        messages.success(request, 'FAQ actualizada correctamente')
    else:
        # Crear
        FAQ.objects.create(pregunta=pregunta, respuesta=respuesta, creado_por=request.user)
        messages.success(request, 'FAQ creada correctamente')

    # --- REDIRECCIÓN CORREGIDA ---
    if es_admin:
        response = redirect('dashboard_admin')
        response['Location'] += '?tab=faq' # Truco para mantener la pestaña
        return response
    else:
        response = redirect('dashboard_tecnico')
        response['Location'] += '?tab=faq' # Lo mismo para el técnico
        return response

@login_required
def eliminar_faq(request, faq_id):
    # Validar permisos (Igual que arriba)
    es_admin = request.user.is_staff
    es_tecnico = False
    try:
        if hasattr(request.user, 'userprofile') and request.user.userprofile.role == 'tech':
            es_tecnico = True
    except:
        pass

    if es_admin or es_tecnico:
        faq = get_object_or_404(FAQ, id=faq_id)
        faq.delete() # O faq.activo = False si prefieres soft-delete
        messages.success(request, 'FAQ eliminada')
    
    if es_admin:
        response = redirect('dashboard_admin')
        response['Location'] += '?tab=faq'
        return response
    
    response = redirect('dashboard_tecnico')
    response['Location'] += '?tab=faq'
    return response

@login_required
def eliminar_categoria(request, categoria_id):
    # Solo staff puede eliminar
    if not request.user.is_staff:
        messages.error(request, 'No tienes permisos para realizar esta acción.')
        return redirect('dashboard_user')

    categoria = get_object_or_404(Categoria, id=categoria_id)
    categoria.delete()
    messages.success(request, f'Categoría "{categoria.nombre}" eliminada correctamente.')

    response = redirect('dashboard_admin')
    response['Location'] += '?tab=categorias'
    return response

@login_required
def cambiar_prioridad_ticket(request, ticket_id):
    # Solo Admin y Técnicos pueden cambiar prioridad
    es_tecnico = False
    try:
        if hasattr(request.user, 'userprofile') and request.user.userprofile.role == 'tech':
            es_tecnico = True
    except:
        pass

    # --- CAMBIO 1: PERMISOS ABIERTOS ---
    # Antes verificábamos si el ticket era del usuario. Ahora solo si es staff o técnico.
    if not (request.user.is_staff or es_tecnico):
        messages.error(request, 'No tienes permiso.')
        return redirect('dashboard_admin')

    ticket = get_object_or_404(Ticket, id=ticket_id)
    
    # Ciclo de prioridades: Baja -> Media -> Alta -> Crítica -> Baja
    prioridades = ['low', 'medium', 'high', 'critical']
    
    try:
        index_actual = prioridades.index(ticket.prioridad)
        nuevo_index = (index_actual + 1) % len(prioridades)
        ticket.prioridad = prioridades[nuevo_index]
        ticket.save()
        
        # Registrar en historial
        registrar_historial(ticket, request.user, f"Prioridad cambiada a {ticket.get_prioridad_display()}")
        messages.success(request, f'Prioridad actualizada a {ticket.get_prioridad_display()}')
    except ValueError:
        # Si hay un error con el valor actual, lo reseteamos a Media
        ticket.prioridad = 'medium'
        ticket.save()

    # --- CAMBIO 2: REDIRECCIÓN CORRECTA ---
    if request.user.is_staff:
        return redirect(reverse('dashboard_admin') + '?tab=tickets')
    else:
        return redirect(reverse('dashboard_tecnico') + '?tab=mis-tickets')

@login_required
@require_POST
def guardar_area(request):
    # Solo staff/admin puede gestionar áreas
    if not request.user.is_staff:
        messages.error(request, 'No tienes permisos.')
        return redirect('dashboard_user')

    try:
        area_id = request.POST.get('area_id')
        nombre = request.POST.get('nombre')
        descripcion = request.POST.get('descripcion')
        
        if area_id:
            # Editar existente
            area = get_object_or_404(Area, id=area_id)
            area.nombre = nombre
            area.descripcion = descripcion
            area.save()
            messages.success(request, f'Área "{nombre}" actualizada correctamente.')
        else:
            # Crear nueva
            Area.objects.create(nombre=nombre, descripcion=descripcion)
            messages.success(request, f'Área "{nombre}" creada correctamente.')
            
    except Exception as e:
        messages.error(request, f'Error al guardar área: {str(e)}')

    # Redirigir manteniendo la pestaña
    response = redirect('dashboard_admin')
    response['Location'] += '?tab=areas'
    return response

@login_required
def exportar_reporte_excel(request):
    # Seguridad: Solo personal autorizado
    if not request.user.is_staff:
        return redirect('dashboard_user')

    # Crear el libro de trabajo
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Reporte de Tickets"

    # --- ESTILOS ---
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="10b981", end_color="10b981", fill_type="solid") # Verde corporativo
    center_align = Alignment(horizontal="center", vertical="center")

    # --- ENCABEZADOS ---
    headers = ["ID", "Título", "Categoría", "Área", "Solicitante", "Técnico", "Prioridad", "Estado", "Fecha Creación", "Satisfacción"]
    ws.append(headers)

    # Aplicar estilo a encabezados
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align

    # --- DATOS ---
    tickets = Ticket.objects.all().order_by('-fecha_creacion')
    for t in tickets:
        # Traducir estado y prioridad
        estado = t.get_estado_display()
        prioridad = t.get_prioridad_display()
        tecnico = t.asignado_a.username if t.asignado_a else "Sin asignar"
        area = t.area.nombre if t.area else "-"
        cat = t.categoria.nombre if t.categoria else "-"
        sat = f"{t.calificacion} ★" if t.calificacion else "-"
        fecha = t.fecha_creacion.strftime('%d/%m/%Y')

        row = [
            f"T-{t.id:04d}", t.titulo, cat, area, 
            t.solicitante.username, tecnico, prioridad, estado, fecha, sat
        ]
        ws.append(row)

    # Ajustar ancho de columnas automáticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except: pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Generar respuesta HTTP
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Tickets_Coyahue_{timezone.now().strftime("%Y-%m-%d")}.xlsx"'
    wb.save(response)
    return response

@login_required
def exportar_reporte_pdf(request):
    if not request.user.is_staff:
        return redirect('dashboard_user')

    # 1. Obtener filtros de la URL (GET)
    f_inicio = request.GET.get('inicio')
    f_fin = request.GET.get('fin')
    f_categoria = request.GET.get('categoria')
    f_prioridad = request.GET.get('prioridad')
    f_tecnico = request.GET.get('tecnico')

    # 2. Filtrar la base de datos
    tickets = Ticket.objects.all().order_by('-fecha_creacion')

    if f_inicio:
        tickets = tickets.filter(fecha_creacion__date__gte=f_inicio)
    if f_fin:
        tickets = tickets.filter(fecha_creacion__date__lte=f_fin)
    
    if f_categoria and f_categoria != "Todas las Categorías":
        tickets = tickets.filter(categoria__nombre=f_categoria)
    
    if f_prioridad and f_prioridad != "Todas las Prioridades":
        # Asegurarnos de mapear el valor del select al valor de la BD si es necesario
        # Si el select envía 'high' y la BD usa 'high', estamos bien.
        tickets = tickets.filter(prioridad=f_prioridad)

    if f_tecnico and f_tecnico != "Todos los Técnicos":
        tickets = tickets.filter(asignado_a__username=f_tecnico)

    # 3. Generar el PDF (Código estándar de ReportLab)
    response = HttpResponse(content_type='application/pdf')
    filename = f'Reporte_Tickets_{timezone.now().strftime("%Y-%m-%d")}.pdf'
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    doc = SimpleDocTemplate(response, pagesize=landscape(letter))
    elements = []
    styles = getSampleStyleSheet()

    # Título con filtros aplicados (Opcional, para que se vea pro)
    titulo_texto = "Reporte de Tickets - Coyahue S.A."
    if f_categoria: titulo_texto += f" ({f_categoria})"
    
    elements.append(Paragraph(titulo_texto, styles['Title']))
    elements.append(Spacer(1, 20))

    # Tabla
    data = [['ID', 'Título', 'Categoría', 'Solicitante', 'Técnico', 'Estado', 'Fecha', 'Calif.']]
    
    for t in tickets:
        titulo_corto = (t.titulo[:25] + '..') if len(t.titulo) > 25 else t.titulo
        tecnico = t.asignado_a.username if t.asignado_a else "-"
        sat = str(t.calificacion) if t.calificacion else "-"
        
        data.append([
            f"T-{t.id:04d}", titulo_corto, t.categoria.nombre, 
            t.solicitante.username, tecnico, t.get_estado_display(), 
            t.fecha_creacion.strftime('%d/%m'), sat
        ])

    t = Table(data)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#10b981')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
    ]))
    
    elements.append(t)
    doc.build(elements)
    return response

@login_required
def eliminar_area(request, area_id):
    if not request.user.is_staff:
        messages.error(request, 'No tienes permisos.')
        return redirect('dashboard_user')
        
    area = get_object_or_404(Area, id=area_id)
    area.delete()
    messages.success(request, 'Área eliminada correctamente.')
    
    response = redirect('dashboard_admin')
    response['Location'] += '?tab=areas'
    return response